from openpyxl import load_workbook
from openpyxl import Workbook


filename = 'rpachallenge.xlsx'

fields_row        = 1
fields_col_start  = 1
fields_col_end    = 7
records_row_start = 2
records_row_end   = 11

wb = load_workbook(filename).active

def extract_fields() -> list[str]:
    fields = []
    for col in range(fields_col_start, fields_col_end+1):
        field = wb.cell(row=fields_row, column=col).value.strip()
        fields.append(field)
    return fields

def extract_records(fields: list[str]) -> list[dict]:
    records = []
    for row in range(records_row_start, records_row_end+1):
        record = {}
        for col in range(fields_col_start, fields_col_end+1):
            field = fields[col-1]
            record[field] = wb.cell(row=row, column=col).value
        records.append(record)
    return records

def main():
    fields = extract_fields()
    print(fields)

if __name__ == '__main__':
    main()
